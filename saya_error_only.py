import re
import fileinput
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font
import sys
import getopt


try:
    opts, args = getopt.getopt(sys.argv[1:], "hi:p:", ["help", "input=", "PSNR="])


except getopt.GetoptError:
    print("Error,please input <inputfile> and <psnr> like \'-i <inputfile> -p <psnr>\'")



for opt,arg in opts:
    if opt == "-h":
        print
        '-i <inputfile> -p <PSNR limit>'
        print
        'or: --infile=<inputfile> --PSNR=<PSNR limit>'

    if opt in ("-i","--input"):
        inputfile=arg

    if opt in ("-p","--PSNR"):
        psnr=arg





text=fileinput.input(inputfile,mode="r")
excel=openpyxl.Workbook()
ws=excel.active
ws.append(["Comparing", "channel(s)", "YUV"])
ws.append(['', '', ''])
ws.append(['', 'Mean', '', 'Max', ''])
ws.append(['', 'Absolute', 'Mean', 'Pos.', 'Neg.'])
ws.append(['Frame', 'Dev.', 'Dev.', 'Dev.', 'Dev.', 'PSNR (dB)'])
'''
# for line in saya:
# #Comparing channel(s) YUV
#     if re.search(r'Comparing\schannel\(s\)\sYUV',line)!=None:
#         str=re.search(r'Comparing\schannel\(s\)\sYUV',line).group()
'''

for line in text:
    ln = text.lineno()

    if ln>6:

        str = re.findall(r'\s*(\+*-*\w+\.*\w*)\s*', line)

        next = re.findall(r'----------\w*\.+\w*----------', line)


        if next!=[]:
            ws.append(next)
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append(['', 'Mean', '', 'Max', ''])
            ws.append(['', 'Absolute', 'Mean', 'Pos.', 'Neg.'])
            ws.append(['Frame', 'Dev.', 'Dev.', 'Dev.', 'Dev.', 'PSNR (dB)'])
            continue

        if len(str) < 6:
            continue
        '''
        #print(ws.max_row)

        # if float(str[1])>#:  #Mean Absolute limit
        #     red=Font(color=colors.RED)
        #     cell=ws["B%d"%(ws.max_row)]
        #     cell.font=red
        #
        # if float(str[2])>#:  #Mean limit
        #     red=Font(color=colors.RED)
        #     cell=ws["C%d"%(ws.max_row)]
        #     cell.font=red
        #
        # if int(str[3])>#:  #Max Pos. limit
        #     red=Font(color=colors.RED)
        #     cell=ws["D%d"%(ws.max_row)]
        #     cell.font=red
        #
        # if int(str[4])>#:  #Max Neg. limit
        #     red=Font(color=colors.RED)
        #     cell=ws["E%d"%(ws.max_row)]
        #     cell.font=red
        '''
        if len(str)==6:
            if str[0]=="Mean":
                continue
            # else:
            #     ws.append(str)

            if float(str[5]) <float(psnr):  # PSNR (dB) limit
                ws.append(str)
                red = Font(color=colors.RED)
                cell = ws["F%d" % (ws.max_row)]
                cell.font = red




excel.save('%s.xlsx'%(inputfile))
